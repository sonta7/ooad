# 课程编码
# 课程名称
# 分组名称
# 开课单位
# 学时
# 学分
# 上课教师
# 上课时间
# 上课地点

import pymysql
import openpyxl

timeSet = set()
weekTypeSet = set()


def excel_operation():
    excel_path = './data/course.xlsx'
    filename = './data/class_info.txt'
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'courses'

    with open(filename, encoding='utf-8') as f:
        lines = f.readlines()
        for i in range(0, len(lines), 9):
            code = lines[i].replace('\n', '')
            classname = lines[i+1].replace('\n', '')
            coursename = lines[i+2].replace('\n', '')
            department = lines[i+3].replace('\n', '')
            period = lines[i+4].replace('\n', '')
            credits = lines[i+5].replace('\n', '')
            teachers = lines[i+6].replace('\n', '')
            times = lines[i+7].replace('\n', '')
            classrooms = lines[i+8].replace('\n', '')
            k = (code, classname, coursename, department, period, credits, teachers, times, classrooms)
            for j in range(9):
                sheet.cell(row=i/9+1, column=j+1, value=k[j])
    wb.save(excel_path)
    print('excel finish')

days = dict((("星期一",'Mon'),('星期二','Tue'),('星期三','Wed'),('星期四','Thu'),('星期五','Fri'),('星期六','Sat'),('星期日','Sun')))
weeks = dict((("1-16周", 0), ('单周', 1), ('双周', 2)))
perioddict = dict((('0102', 1),('0304', 2),('0506', 3),('0708', 4),('0910', 5)))

class Course():
    def __init__(self, code, classname, coursename, department, period, credits, teachers, times, classrooms):
        self.code = code
        self.classname = classname
        self.coursename = classname +' '+ coursename
        self.department = department
        self.period = period
        self.credits = credits

        if teachers == None:
            teacherlist = []
        else:
            teacherlist = teachers.split(',')
        self.teachers = teacherlist
        
        self.classtimes = []
        timelist = times.split(' ')
        if len(timelist) != 1:
            for i in range(0, len(timelist), 3):
                weekType = weeks[timelist[i]]
                day = days[timelist[i+1]]
                
                periods = timelist[i+2]


                weekTypeSet.add(weekType)

                # if weekType != '1-16'
                for j in range(0, len(periods)-1, 4):
                    period_j = periods[j:j+4]
                    periodid = perioddict[period_j]
                    # print(periodid)
                    # if period_j == '0102':
                    #     periodid = 1
                    # elif period_j == '0304':
                    #     periodid = 2
                    # elif period_j == '0506':
                    #     periodid = 3
                    # elif period_j == '0708':
                    #     periodid = 4
                    # elif period_j == '0910':
                    #     periodid = 5


                    timeSet.add((day, periodid))
                    self.classtimes.append((weekType, day, periodid))

        if classrooms == None:
            classroomlist = []
        else:
            classroomlist = classrooms.split(' ')
        self.classrooms = classroomlist
        for _ in range(len(self.classrooms), len(self.classtimes)):
            self.classrooms.append('')
        self.time_room = []
        if len(self.classrooms)>1:
            for i in range(len(self.classrooms)):
                self.time_room.append((self.classtimes[i], self.classrooms[i]))

def file_reader(filename):
    classSet = set()
    departmentSet = set()
    teacherSet = set()
    classroomSet = set()
    courseList = []
    with open(filename, encoding='utf-8') as f:
        lines = f.readlines()
        for i in range(0, len(lines), 9):
            code = lines[i].replace('\n', '')
            classname = lines[i+1].replace('\n', '')
            coursename = lines[i+2].replace('\n', '')
            department = lines[i+3].replace('\n', '')
            period = lines[i+4].replace('\n', '')
            credits = lines[i+5].replace('\n', '')
            teachers = lines[i+6].replace('\n', '')
            times = lines[i+7].replace('\n', '')
            classrooms = lines[i+8].replace('\n', '')
            course = Course(code, classname, coursename, department, period, credits, teachers, times, classrooms)
            
            courseList.append(course)
            
            classSet.add((code, classname, period, credits, department))
            departmentSet.add(department)
            
            teacherlist = teachers.split(',')
            for t in teacherlist:
                teacherSet.add(t)
            
            classroomlist = classrooms.split(' ')
            for room in classroomlist:
                classroomSet.add(room)
    
    return classSet, departmentSet, teacherSet, classroomSet, courseList, weekTypeSet, timeSet

# excel_operation()

# classSet, departmentSet, teacherSet, classroomSet, courseList, weekTypeSet, timeSet= file_reader('./data/class_info.txt')
# classList = sorted(list(classSet))

def excel_read(filename='./data/course.xlsx'):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    classSet = set()
    departmentSet = set()
    teacherSet = set()
    classroomSet = set()
    courseList = []
    
    for row in ws.rows:
        code = row[0].value
        classname = row[1].value
        coursename = row[2].value
        department = row[3].value
        period = row[4].value
        credits = row[5].value
        teachers = row[6].value
        times = row[7].value
        classrooms = row[8].value
        course = Course(code, classname, coursename, department, period, credits, teachers, times, classrooms)
        
        courseList.append(course)
        
        classSet.add((code, classname, period, credits, department))
        departmentSet.add(department)
        
        teacherlist = teachers.split(',') if teachers is not None else []
        for t in teacherlist:
            teacherSet.add(t)
        
        classroomlist = classrooms.split(' ') if classrooms is not None else []
        for room in classroomlist:
            classroomSet.add(room)
    
    return classSet, departmentSet, teacherSet, classroomSet, courseList, weekTypeSet, timeSet


# classSet, departmentSet, teacherSet, classroomSet, courseList, weekTypeSet, timeSet = excel_read()  
# print(weekTypeSet)
# print(departmentSet)

# print(classList, '\n')
# print(departmentSet, '\n')
# print(teacherSet, '\n')
# print(sorted(list(classroomSet)), '\n')
# print(sorted(weekTypeSet), '\n')
# print(sorted(timeSet), '\n')

# print("finish")
